import os
import json
from pypdf import PdfReader
import openpyxl

DATA_DIR = os.path.join(os.path.dirname(__file__), "data")
OUTPUT_FILE = os.path.join(os.path.dirname(__file__), "raw_docs.json")

RELEVANT_FILES = [
    "1.1_Statuts_FallahTech_1775661590518.pdf",
    "1.2_Contrat_Cooperative_Type_1775661590518.pdf",
    "2.1_Etats_Financiers_Historiques_NCT_2023_2025_1775661590519.pdf",
    "3.1_Registre_Personnel_1775661590519.pdf",
    "4.1_Etude_Marche_Synthese_1775661590519.pdf",
    "FallahTech_BusinessPlan_Complet_1775661590520.xlsx",
]


def extract_pdf_text(filepath):
    reader = PdfReader(filepath)
    pages = []
    for i, page in enumerate(reader.pages):
        text = page.extract_text() or ""
        text = text.replace("\x00", "")
        if text.strip():
            pages.append({"page": i + 1, "text": text.strip()})
    return pages


def extract_xlsx_text(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheets = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            row_text = " | ".join([str(c) if c is not None else "" for c in row])
            if row_text.strip().replace("|", "").strip():
                rows.append(row_text)
        if rows:
            sheets.append({"sheet": sheet_name, "text": "\n".join(rows)})
    return sheets


def ingest_documents():
    documents = []
    for filename in RELEVANT_FILES:
        filepath = os.path.join(DATA_DIR, filename)
        if not os.path.exists(filepath):
            continue

        source_name = filename.split("_1775661")[0] if "_1775661" in filename else filename
        if filename.endswith(".pdf"):
            pages = extract_pdf_text(filepath)
            for page_data in pages:
                documents.append({
                    "source": source_name + ".pdf",
                    "page": page_data["page"],
                    "text": page_data["text"],
                })
        elif filename.endswith(".xlsx"):
            sheets = extract_xlsx_text(filepath)
            for sheet_data in sheets:
                documents.append({
                    "source": source_name + ".xlsx",
                    "sheet": sheet_data["sheet"],
                    "text": sheet_data["text"],
                })

    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(documents, f, ensure_ascii=False, indent=2)

    return documents


if __name__ == "__main__":
    docs = ingest_documents()
    print(f"Ingested {len(docs)} document segments.")
